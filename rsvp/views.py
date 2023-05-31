from rest_framework.authentication import SessionAuthentication, BasicAuthentication
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Permission

from .emails.main import EmailClient
from django.shortcuts import render
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.hashers import check_password
from django.conf import settings

import json

# Spreadsheets
from functools import cmp_to_key
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder


from django.contrib.auth.models import User, Group
from rest_framework import viewsets, mixins
from rsvp.serializers import InviteSerializer
from .models import Invite, Guest, Email


class InviteViewSet(mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    queryset = Invite.objects.all().order_by('family_name')
    serializer_class = InviteSerializer


@csrf_exempt
def submit_rsvp(request):
    if request.method != "POST":
        response = {"status": "error", "message": "Method not allowed"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=405)
    body_unicode = request.body.decode('utf-8')
    body = json.loads(body_unicode)
    if not "uuid" in body or not body['uuid']:
        response = {"status": "error",
                    "message": "an invite UUID was not provided"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=400)
    try:
        invite = Invite.objects.filter(uuid=body['uuid']).first()
    except ValidationError:
        response = {"status": "error",
                    "message": "the UUID provided is not valid"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=400)

    if not invite:
        response = {"status": "error",
                    "message": "an invite with the provided UUID was not found"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=404)

    if not "guests" in body or len(body['guests']) != len(invite.guests.all()):
        response = {"status": "error",
                    "message": "the guests array was not provided or is invalid"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=400)

    guests = body['guests']
    for guest in guests:
        guest_obj = Guest.objects.filter(id=guest['id']).first()
        if not guest_obj:
            response = {"status": "error",
                        "message": f"the guest with ID {guest['id']} was not found."}
            return HttpResponse(json.dumps(response), content_type="application/json", status=404)
        guest_obj.is_attending = guest['is_attending']
        print(guest["is_attending"])
        guest_obj.dietary_restrictions = guest['dietary_restrictions'] if guest_obj.is_attending else ""
        guest_obj.save()

    invite.finished = True
    invite.save()
    response = {"status": "ok",
                "message": "your RSVP was successfully saved!"}
    return HttpResponse(json.dumps(response), content_type="application/json", status=200)


@csrf_exempt
def send_emails(request):
    if request.method != "POST":
        response = {"status": "error", "message": "Method not allowed"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=405)

    body_unicode = request.body.decode('utf-8')
    body = json.loads(body_unicode)

    if not "username" in body or not "password" in body:
        response = {"status": "error",
                    "message": "Username or password was not provided"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=400)

    user = User.objects.filter(username=body['username']).first()

    if not user:
        response = {"status": "error",
                    "message": "a user with this username was not found"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=404)

    if not user.check_password(body['password']):
        response = {"status": "error",
                    "message": "the password provided was not correct"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=401)

    if not user.has_perms(["superuser"]):
        response = {"status": "error",
                    "message": "the user does not have superuser permissions"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=401)

    if not "email_content" in body:
        response = {"status": "error",
                    "message": "an email body was not provided as \"email_content\""}
        return HttpResponse(json.dumps(response), content_type="application/json", status=400)

    if not "subject" in body:
        response = {"status": "error",
                    "message": "an email subject was not provided as \"subject\""}
        return HttpResponse(json.dumps(response), content_type="application/json", status=400)

    content: str = body['email_content']

    all_emails = "all_emails" in body and body['all_emails'] == True

    emails = Email.objects.all()
    if not all_emails:
        emails = emails.filter(invite__finished=False)

    client = EmailClient()
    for email in emails:
        client.send_email(
            email.address, body['subject'], content.format(**email.invite.__dict__))

    response = {"status": "ok",
                "message": f"Successfully sent {emails.__len__()} emails!"}
    return HttpResponse(json.dumps(response), content_type="application/json", status=200)


@csrf_exempt
def send_specific(request):
    if request.method != "POST":
        response = {"status": "error", "message": "Method not allowed"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=405)

    body_unicode = request.body.decode('utf-8')
    body = json.loads(body_unicode)

    if not "username" in body or not "password" in body:
        response = {"status": "error",
                    "message": "Username or password was not provided"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=400)

    user = User.objects.filter(username=body['username']).first()

    if not user:
        response = {"status": "error",
                    "message": "a user with this username was not found"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=404)

    if not user.check_password(body['password']):
        response = {"status": "error",
                    "message": "the password provided was not correct"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=401)

    if not user.has_perms(["superuser"]):
        response = {"status": "error",
                    "message": "the user does not have superuser permissions"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=401)

    if not "email_content" in body:
        response = {"status": "error",
                    "message": "an email body was not provided as \"email_content\""}
        return HttpResponse(json.dumps(response), content_type="application/json", status=400)

    if not "subject" in body:
        response = {"status": "error",
                    "message": "an email subject was not provided as \"subject\""}
        return HttpResponse(json.dumps(response), content_type="application/json", status=400)

    if not "rowSelection" in body:
        response = {"status": "error",
                    "message": "a row selection array was not provided"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=400)

    content: str = body['email_content']

    all_emails = "all_emails" in body and body['all_emails'] == True

    emails = [Email.objects.filter(id=_id).first()
              for _id in body['rowSelection']]

    client = EmailClient()
    for email in emails:
        guests = [guest for guest in Guest.objects.filter(
            invite=email.invite, is_child=False)]
        if len(guests) == 1:
            guest_names = f"{guests[0].name}"
        elif len(guests) == 2:
            guest_names = f"{guests[0]} and {guests[1]}"
        else:
            guest_names = f"{', '.join([guest.name for guest in guests[:-1]])}, and {guests[-1].name}"

        client.send_email(
            email.address, body['subject'], content.format(**email.invite.__dict__, guest_names=guest_names))

    response = {"status": "ok",
                "message": f"Successfully sent {emails.__len__()} emails!"}
    return HttpResponse(json.dumps(response), content_type="application/json", status=200)


@csrf_exempt
def test_email(request):
    if request.method != "POST":
        response = {"status": "error", "message": "Method not allowed"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=405)

    body_unicode = request.body.decode('utf-8')
    body = json.loads(body_unicode)

    if not "username" in body or not "password" in body:
        response = {"status": "error",
                    "message": "Username or password was not provided"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=400)

    user = User.objects.filter(username=body['username']).first()

    if not user:
        response = {"status": "error",
                    "message": "a user with this username was not found"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=404)

    if not user.check_password(body['password']):
        response = {"status": "error",
                    "message": "the password provided was not correct"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=401)

    if not user.has_perms(["superuser"]):
        response = {"status": "error",
                    "message": "the user does not have superuser permissions"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=401)

    if not "email_content" in body:
        response = {"status": "error",
                    "message": "an email body was not provided as \"email_content\""}
        return HttpResponse(json.dumps(response), content_type="application/json", status=400)

    if not "subject" in body:
        response = {"status": "error",
                    "message": "an email subject was not provided as \"subject\""}
        return HttpResponse(json.dumps(response), content_type="application/json", status=400)

    content: str = body['email_content']

    client = EmailClient()
    email = Email.objects.first()

    if not email:
        response = {"status": "error",
                    "message": "there are no emails in the database yet"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=404)

    client.send_email(
        client.sender, body['subject'], content.format(**email.invite.__dict__))

    response = {"status": "ok",
                "message": f"Successfully sent test email!"}
    return HttpResponse(json.dumps(response), content_type="application/json", status=200)


@csrf_exempt
def dry_run(request):
    if request.method != "POST":
        response = {"status": "error", "message": "Method not allowed"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=405)

    body_unicode = request.body.decode('utf-8')
    body = json.loads(body_unicode)

    if not "username" in body or not "password" in body:
        response = {"status": "error",
                    "message": "Username or password was not provided"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=400)

    user = User.objects.filter(username=body['username']).first()

    if not user:
        response = {"status": "error",
                    "message": "a user with this username was not found"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=404)

    if not user.check_password(body['password']):
        response = {"status": "error",
                    "message": "the password provided was not correct"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=401)

    if not user.has_perms(["superuser"]):
        response = {"status": "error",
                    "message": "the user does not have superuser permissions"}
        return HttpResponse(json.dumps(response), content_type="application/json", status=401)

    all_emails = "all_emails" in body and body['all_emails'] == True

    emails = Email.objects.all()
    if not all_emails:
        emails = emails.filter(invite__finished=False)

    response = {"status": "ok",
                "message": f"Fetched {emails.__len__()} emails!", "emails": [{"id": email.id, "address": email.address, "family_name": email.invite.family_name} for email in emails]}
    return HttpResponse(json.dumps(response), content_type="application/json", status=200)


@login_required
def generate_xlsx_spreadsheet(request):
    filename = "melanie_andrew_wedding_rsvp.xlsx"

    def guest_sorter(guest_1: Guest, guest_2: Guest):
        if guest_1.invite.finished != guest_2.invite.finished:
            return guest_2.invite.finished - guest_1.invite.finished
        if guest_1.is_attending != guest_2.is_attending:
            return guest_2.is_attending - guest_1.is_attending
        return int(guest_1.invite.family_name < guest_2.invite.family_name)

    guests = [_ for _ in Guest.objects.all()]
    guests.sort(key=cmp_to_key(guest_sorter))

    """
    Generate Excel spreadsheet
    """
    # Header
    data = [["GUEST NAME", "FAMILY NAME",
             "IS ATTENDING?", "DIETARY RESTRICTIONS"]]

    # Raw Data
    for guest in guests:
        data.append([guest.name, guest.invite.family_name, ("Yes" if guest.is_attending else "No",
                    guest.dietary_restrictions) if guest.invite.finished else "Incomplete RSVP"])

    # Analytics
    data.append([])
    data.append([])
    data.append([])
    data.append(["Number of people confirmed attending:",
                Guest.objects.filter(is_attending=True).__len__()])
    data.append(["Number of people confirmed NOT attending:",
                Guest.objects.filter(is_attending=False).__len__()])
    data.append(["Total guests who completed RSVP:",
                Guest.objects.filter(invite__finished=True).__len__()])
    data.append(["Total guests invited:", Guest.objects.all().__len__()])

    wb = Workbook()
    ws = wb.active
    ws.title = "Guest List"

    for row in data:
        ws.append(row)

    dim_holder = DimensionHolder(worksheet=ws)

    # https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
    # for col in range(ws.min_column, ws.max_column + 1):
    #     dim_holder[get_column_letter(col)] = ColumnDimension(
    #         ws, min=col, max=col, width=20)

    dim_holder['A'] = ColumnDimension(ws, min=1, max=1, width=43)
    dim_holder['B'] = ColumnDimension(ws, min=2, max=2, width=43)
    dim_holder['C'] = ColumnDimension(ws, min=3, max=3, width=20)
    dim_holder['D'] = ColumnDimension(ws, min=4, max=4, width=100)
    ws.column_dimensions = dim_holder

    ft = Font(bold=True)
    for row in ws['A1:D1']:
        for cell in row:
            cell.font = ft

    wb.save(filename)

    # Return it to the user
    with open(filename, 'rb') as f:
        file_data = f.read()

    response = HttpResponse(
        file_data, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="melanie_andrew_wedding_rsvp.xlsx"'

    return response

def rsvp_deadline(request):
    response = {"status": "ok", "deadline": settings.RSVP_DEADLINE.isoformat()}
    return HttpResponse(json.dumps(response), content_type="application/json", status=200)